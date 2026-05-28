"""
build_workpaper.py
------------------
Builds a formatted Excel workpaper for non-technical analysts.

Run after the pipeline:
    python build_workpaper.py

Output:
    outputs/fraud_risk_workpaper.xlsx
"""

from __future__ import annotations

import json
import shutil
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import config

# ── Palette (aligned with EDA / dashboard) ────────────────────────────────────
NAVY = "1B2A4A"
WHITE = "FFFFFF"
HIGH_FILL = "FADBD8"
HIGH_FONT = "922B21"
MED_FILL = "FDEBD0"
MED_FONT = "B7950B"
LOW_FILL = "D5F5E3"
LOW_FONT = "1E8449"
FRAUD_FILL = "FCF3CF"
HEADER_FILL = "1B2A4A"
ALT_ROW = "F8F9FA"
BORDER_COLOR = "D5D8DC"

WORKPAPER_PATH = config.WORKPAPER_XLSX

REVIEW_STATUSES = '"Pending,In Review,Escalated,Cleared"'

# Slightly larger fonts for analyst readability (avoid Excel Table objects — they corrupt styled sheets)
FONT_TITLE = 18
FONT_SECTION = 13
FONT_HEADER = 12
FONT_BODY = 12
FONT_CELL = 11


def _load_optional_csv(path: Path) -> pd.DataFrame | None:
    if path.exists():
        return pd.read_csv(path)
    return None


def _load_topic_meta() -> dict:
    path = config.SAR_TOPIC_META_JSON
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {"z_threshold": 1.0, "n_topics": 5, "stop_words": []}


def _style_risk_score_column(ws, headers: list[str], nrows: int):
    """Conditional fill on risk_score (0–100)."""
    if "risk_score" not in headers or nrows < 1:
        return
    col = get_column_letter(headers.index("risk_score") + 1)
    from openpyxl.formatting.rule import ColorScaleRule

    ws.conditional_formatting.add(
        f"{col}2:{col}{nrows + 1}",
        ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="E6F4F1",
            mid_type="num",
            mid_value=50,
            mid_color="FDEBD0",
            end_type="num",
            end_value=100,
            end_color="FADBD8",
        ),
    )


def _metrics(scored: pd.DataFrame, raw: pd.DataFrame | None) -> dict:
    n = len(scored)
    tiers = scored["risk_tier"].value_counts() if "risk_tier" in scored.columns else {}
    fraud_n = int(scored["isFraud"].sum()) if "isFraud" in scored.columns else None
    high_fraud = None
    if fraud_n is not None and "risk_tier" in scored.columns:
        high_fraud = int(
            scored[(scored["isFraud"] == 1) & (scored["risk_tier"] == "HIGH")].shape[0]
        )
    return {
        "total_transactions": n,
        "raw_transactions": len(raw) if raw is not None else n,
        "flagged_anomalies": int(scored["anomaly_flag"].sum()) if "anomaly_flag" in scored.columns else None,
        "high_count": int(tiers.get("HIGH", 0)),
        "medium_count": int(tiers.get("MEDIUM", 0)),
        "low_count": int(tiers.get("LOW", 0)),
        "avg_risk_score": round(scored["risk_score"].mean(), 1) if "risk_score" in scored.columns else None,
        "labeled_fraud": fraud_n,
        "fraud_in_high": high_fraud,
        "fraud_rate_pct": round(100 * fraud_n / n, 2) if fraud_n is not None and n else None,
    }


def _thin_border():
    side = Side(style="thin", color=BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)


def _sanitize_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Clean dtypes and column names so openpyxl/Excel do not mis-read cells."""
    if df is None or df.empty:
        return df
    out = df.copy()
    # Unique column names (duplicate headers break Excel)
    seen: dict[str, int] = {}
    new_cols = []
    for c in out.columns:
        name = str(c).strip()
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        new_cols.append(name)
    out.columns = new_cols

    for col in out.columns:
        series = out[col]
        if pd.api.types.is_bool_dtype(series):
            out[col] = series.map({True: "Yes", False: "No", np.True_: "Yes", np.False_: "No"}).fillna("")
        elif pd.api.types.is_numeric_dtype(series):
            out[col] = series.replace([np.inf, -np.inf], np.nan)
        else:
            out[col] = series.fillna("").astype(str).replace({"nan": "", "None": ""})
    return out


def _cell_value(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    if isinstance(val, (np.integer,)):
        return int(val)
    if isinstance(val, (np.floating,)):
        return float(val) if not pd.isna(val) else ""
    if isinstance(val, (np.bool_, bool)):
        return "Yes" if val else "No"
    return val


def _run_date_range(scored: pd.DataFrame, raw: pd.DataFrame | None) -> str:
    if "step" in scored.columns and scored["step"].notna().any():
        lo, hi = int(scored["step"].min()), int(scored["step"].max())
        return f"Transaction steps {lo:,} through {hi:,} (synthetic timeline index)"
    if raw is not None and "step" in raw.columns and raw["step"].notna().any():
        lo, hi = int(raw["step"].min()), int(raw["step"].max())
        return f"Source transaction steps {lo:,} through {hi:,}"
    mtime = config.SCORED_CSV.stat().st_mtime if config.SCORED_CSV.exists() else None
    if mtime:
        return f"Pipeline output dated {datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M')}"
    return f"Run generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"


def _conclusion_insights(metrics: dict, scored: pd.DataFrame) -> list[str]:
    n = metrics["total_transactions"] or 1
    lines = [
        f"HIGH-tier queue: {metrics['high_count']:,} transactions "
        f"({round(100 * metrics['high_count'] / n, 1)}% of the scored population).",
        f"MEDIUM-tier queue: {metrics['medium_count']:,}; LOW-tier: {metrics['low_count']:,}.",
    ]
    if metrics.get("avg_risk_score") is not None:
        lines.append(f"Average risk score across the population: {metrics['avg_risk_score']}.")
    if metrics.get("flagged_anomalies") is not None:
        lines.append(
            f"Isolation Forest flagged {metrics['flagged_anomalies']:,} statistical anomalies "
            f"(target ~{config.ISOLATION_FOREST['contamination'] * 100:.0f}% of volume)."
        )
    if metrics.get("labeled_fraud") is not None and metrics["labeled_fraud"]:
        lines.append(
            f"Labeled fraud (research only): {metrics['labeled_fraud']:,} cases "
            f"({metrics['fraud_rate_pct']}% of population); "
            f"{metrics.get('fraud_in_high') or 0:,} of those appear in the HIGH tier."
        )
    if "country" in scored.columns and "risk_tier" in scored.columns:
        high = scored[scored["risk_tier"] == "HIGH"]
        if len(high):
            top_c = high.groupby("country").size().idxmax()
            lines.append(f"Most HIGH-tier flags by count originate from country code {top_c}.")
    if "amount" in scored.columns and len(scored):
        high_vol = scored.loc[scored["risk_tier"] == "HIGH", "amount"].sum() if "risk_tier" in scored.columns else 0
        tot = scored["amount"].sum()
        if tot:
            lines.append(
                f"HIGH-tier exposure is ${high_vol:,.0f} ({round(100 * high_vol / tot, 1)}% of total volume)."
            )
    lines.append(
        "Prioritize HIGH Risk Alerts and SAR Narratives tabs first; use Risk Register reconciliation "
        "columns before pre-filling case systems."
    )
    return lines


def _style_header_row(ws, row: int, ncol: int):
    fill = PatternFill("solid", fgColor=HEADER_FILL)
    font = Font(bold=True, color=WHITE, size=FONT_HEADER)
    border = _thin_border()
    for col in range(1, ncol + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def _autosize_columns(ws, max_width=48):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        length = max(len(str(c.value or "")) for c in col_cells[:200])
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def _write_dataframe_sheet(
    wb: Workbook,
    title: str,
    df: pd.DataFrame,
    *,
    tier_column: str | None = "risk_tier",
    freeze_row: int = 2,
    extra_columns: list[str] | None = None,
):
    """Write a data sheet with headers, filters, and tier coloring (no Excel Table objects)."""
    if df is None or df.empty:
        ws = wb.create_sheet(title[:31])
        ws["A1"] = f"No records available for '{title}'."
        return ws

    out = _sanitize_for_excel(df.copy())
    if extra_columns:
        for col in extra_columns:
            if col not in out.columns:
                out[col] = ""

    ws = wb.create_sheet(title[:31])
    headers = list(out.columns)
    ncol = len(headers)

    for c, name in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=name)
    _style_header_row(ws, 1, ncol)

    tier_idx = headers.index(tier_column) + 1 if tier_column and tier_column in headers else None

    for r, row in enumerate(out.itertuples(index=False), 2):
        row_fill = ALT_ROW if r % 2 == 0 else WHITE
        tier_val = getattr(row, tier_column, None) if tier_column and tier_column in headers else None
        if tier_val == "HIGH":
            row_fill = HIGH_FILL
        elif tier_val == "MEDIUM":
            row_fill = MED_FILL
        elif tier_val == "LOW":
            row_fill = LOW_FILL

        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=_cell_value(val))
            cell.fill = PatternFill("solid", fgColor=row_fill)
            cell.border = _thin_border()
            cell.font = Font(size=FONT_CELL)
            wrap = headers[c - 1] in ("sar_narrative", "extracted_risk_factors", "topic_top_terms", "top_terms")
            cell.alignment = Alignment(vertical="top", wrap_text=wrap)
            if headers[c - 1] == "isFraud" and val in (1, "1", "Yes"):
                cell.fill = PatternFill("solid", fgColor=FRAUD_FILL)
                cell.font = Font(bold=True, size=FONT_CELL)

    ws.freeze_panes = ws.cell(row=freeze_row, column=1)
    if len(out) > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{len(out) + 1}"

    _autosize_columns(ws)
    if "sar_narrative" in headers:
        letter = get_column_letter(headers.index("sar_narrative") + 1)
        ws.column_dimensions[letter].width = 60
    if "risk_score" in headers:
        _style_risk_score_column(ws, headers, len(out))

    # Review status dropdown on analyst sheets
    if "analyst_review_status" in headers:
        col_letter = get_column_letter(headers.index("analyst_review_status") + 1)
        dv = DataValidation(type="list", formula1=REVIEW_STATUSES, allow_blank=True)
        dv.error = "Choose a status from the list."
        dv.prompt = "Select review disposition."
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{len(out) + 1}")

    return ws


def _write_methodology(
    wb: Workbook,
    metrics: dict,
    paths: dict[str, str],
    *,
    date_range: str,
    insights: list[str],
):
    ws = wb.create_sheet("Methodology", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 82

    title_font = Font(bold=True, size=FONT_TITLE, color=NAVY)
    section_font = Font(bold=True, size=FONT_SECTION, color=NAVY)
    body_font = Font(size=FONT_BODY)
    wrap = Alignment(wrap_text=True, vertical="top")

    def section(row: int, heading: str, lines: list[str]) -> int:
        c = ws.cell(row=row, column=1, value=heading)
        c.font = section_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1
        for line in lines:
            cell = ws.cell(row=row, column=1, value=line)
            cell.font = body_font
            cell.alignment = wrap
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws.row_dimensions[row].height = max(18, min(120, 14 * (1 + len(line) // 90)))
            row += 1
        return row + 1

    row = 1
    t = ws.cell(row=row, column=1, value="Risk Analysis Profile — Analyst Workpaper")
    t.font = title_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 2

    row = section(
        row,
        "Date range for this run",
        [
            date_range,
            f"Workpaper file generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        ],
    )

    row = section(
        row,
        "Purpose",
        [
            "Provide a single, audit-ready workbook for triaging unusual transactions, "
            "reviewing dollar exposure, reading SAR-style narratives, and documenting analyst disposition.",
            "This file is for human review—not model training or automated filing.",
        ],
    )

    row = section(
        row,
        "Source",
        [
            f"Scored transactions: {paths.get('scored', 'N/A')}",
            f"Raw transactions: {paths.get('raw', 'N/A')}",
            f"SAR narratives: {paths.get('narratives', 'N/A')}",
            f"Risk register: {paths.get('register', 'N/A')}",
            f"Topic model outputs: {paths.get('topics', 'N/A')}",
            "Data may be synthetic PaySim-style rows or Kaggle PaySim placed in data/transactions.csv.",
        ],
    )

    row = section(
        row,
        "Scope",
        [
            f"Population reviewed: {metrics['total_transactions']:,} scored transactions.",
            f"HIGH tier (score ≥ {config.RISK_TIER_HIGH}): {metrics['high_count']:,} | "
            f"MEDIUM (≥ {config.RISK_TIER_MEDIUM}): {metrics['medium_count']:,} | "
            f"LOW: {metrics['low_count']:,}.",
            "Includes financial rollups, alert queues, SAR text, NLP risk register, topic themes, and population summaries.",
            "Does not include live core banking balances, customer KYC packets, or filed FinCEN SAR XML.",
        ],
    )

    row = section(
        row,
        "What each tab contains",
        [
            "Methodology (this tab) — Run context, sources, scope, and conclusions.",
            "Executive Summary — Headline metrics and analyst notes.",
            "Financial Summary / Profit by Tier / Profit by Type / Top Account Exposure — Dollar impact views.",
            "HIGH Risk Alerts — Priority queue; set analyst_review_status and analyst_notes.",
            "MEDIUM Risk Alerts — Secondary queue for scheduled review.",
            "SAR Narratives — Investigation memos for HIGH-tier flags.",
            "SAR Topic Summary / SAR Topic Assignments / SAR Word Tree — Narrative theme model from the dashboard.",
            "Topic Model Settings — Z-score, topic count, and stop words used for the theme run.",
            "Risk Register — Structured entities extracted from narratives.",
            "Risk Indicators — Yes/no feature flags in plain language.",
            "Population by Tier / Population by Type — Count and fraud-rate rollups.",
            "Data Dictionary — Field definitions.",
        ],
    )

    row = section(row, "Conclusion — major insights for this run", insights)

    row = section(
        row,
        "Important limitations",
        [
            "Demonstration data only unless you supplied verified institution files.",
            "Risk score measures unusualness, not proof of fraud or SAR filing requirement.",
            "SAR narratives are template-generated for workflow demo unless replaced with real case notes.",
            "Close this workbook in Excel before regenerating from the dashboard to avoid file-lock errors.",
        ],
    )

    return ws


def _write_executive_summary(wb: Workbook, metrics: dict):
    ws = wb.create_sheet("Executive Summary")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 50

    rows = [
        ("Metric", "Value", "Analyst note"),
        ("Total transactions reviewed", metrics["total_transactions"], "Full scored population in pipeline output."),
        ("HIGH risk tier", metrics["high_count"], "Review first—see HIGH Risk Alerts tab."),
        ("MEDIUM risk tier", metrics["medium_count"], "Schedule secondary review."),
        ("LOW risk tier", metrics["low_count"], "Generally monitor only."),
        ("Isolation Forest flags (anomaly_flag=1)", metrics["flagged_anomalies"], f"Target ~{config.ISOLATION_FOREST['contamination']*100:.0f}% of volume."),
        ("Average risk score", metrics["avg_risk_score"], "0–100 scale; higher = more unusual."),
    ]
    if metrics.get("labeled_fraud") is not None:
        rows.append(("Labeled fraud (if present)", metrics["labeled_fraud"], "For model validation only—not used to train the detector."))
        rows.append(
            ("Labeled fraud in HIGH tier", metrics["fraud_in_high"],
             "Shows how well HIGH tier captures known fraud in synthetic/research data."),
        )

    for r, row in enumerate(rows, 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if r > 1:
                cell.font = Font(size=FONT_CELL)
    _style_header_row(ws, 1, 3)
    for r in range(2, len(rows) + 1):
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = _thin_border()
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")


def _write_data_dictionary(wb: Workbook):
    ws = wb.create_sheet("Data Dictionary")
    entries = [
        ("Field", "Description", "Tab(s)"),
        ("transaction_id", "Unique transaction identifier.", "Alerts, SAR"),
        ("risk_score", "0–100 anomaly score; higher = more unusual vs. peers.", "All scored tabs"),
        ("risk_tier", "HIGH / MEDIUM / LOW triage label.", "All scored tabs"),
        ("anomaly_flag", "1 if Isolation Forest flagged as top ~2% anomalous.", "Alerts"),
        ("type", "Transaction type (TRANSFER, CASH_OUT, PAYMENT, etc.).", "Alerts"),
        ("amount", "Transaction dollar amount.", "Alerts"),
        ("country", "Origin country code.", "Alerts"),
        ("hour", "Hour of day (0–23).", "Alerts"),
        ("balance_drained", "1 if origin balance went to zero after transaction.", "Indicators"),
        ("is_late_night", "1 if hour 0–4 (off-hours).", "Indicators"),
        ("is_high_risk_country", "1 if NG, RU, UA, or KE.", "Indicators"),
        ("is_large_transaction", f"1 if amount > ${config.LARGE_TRANSACTION_THRESHOLD:,}.", "Indicators"),
        ("isFraud", "Research fraud label when available—not used for scoring.", "Alerts"),
        ("sar_id", "Synthetic SAR reference ID.", "SAR, Risk Register"),
        ("sar_narrative", "Investigation-style memo text.", "SAR"),
        ("extracted_*", "GLiNER-extracted entities from narrative.", "Risk Register"),
        ("amount_reconciled", "True if structured amount appears in extraction/narrative.", "Risk Register"),
        ("analyst_review_status", "Your disposition: Pending / In Review / Escalated / Cleared.", "Alerts"),
        ("analyst_notes", "Free-text review notes.", "Alerts"),
    ]
    for r, row in enumerate(entries, 1):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    _style_header_row(ws, 1, 3)
    _autosize_columns(ws, max_width=55)


def _indicator_frame(df: pd.DataFrame) -> pd.DataFrame:
    """Human-readable yes/no risk indicator columns for analysts."""
    cols = {
        "transaction_id": df.get("transaction_id"),
        "risk_score": df.get("risk_score"),
        "risk_tier": df.get("risk_tier"),
        "type": df.get("type"),
        "amount": df.get("amount"),
        "country": df.get("country"),
    }
    out = pd.DataFrame(cols)
    flag_map = {
        "Balance fully drained?": "balance_drained",
        "Late night (0–4 AM)?": "is_late_night",
        "High-risk country?": "is_high_risk_country",
        "Large transaction (>$100K)?": "is_large_transaction",
        "High-risk type (TRANSFER/CASH_OUT)?": "is_high_risk_type",
        "Statistical anomaly flag?": "anomaly_flag",
    }
    for label, col in flag_map.items():
        if col in df.columns:
            out[label] = df[col].map({1: "YES", 0: "NO"}).fillna("NO")
    if "isFraud" in df.columns:
        out["Labeled fraud (research)?"] = df["isFraud"].map({1: "YES", 0: "NO"})
    out["analyst_review_status"] = "Pending"
    out["analyst_notes"] = ""
    return out


def _population_summary(scored: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for tier in ["HIGH", "MEDIUM", "LOW"]:
        sub = scored[scored["risk_tier"] == tier] if "risk_tier" in scored.columns else scored.iloc[0:0]
        fraud_rate = None
        if "isFraud" in sub.columns and len(sub):
            fraud_rate = round(100 * sub["isFraud"].mean(), 2)
        rows.append({
            "risk_tier": tier,
            "transaction_count": len(sub),
            "avg_risk_score": round(sub["risk_score"].mean(), 1) if len(sub) and "risk_score" in sub.columns else None,
            "total_amount": round(sub["amount"].sum(), 2) if len(sub) and "amount" in sub.columns else None,
            "fraud_rate_pct": fraud_rate,
            "anomaly_flags": int(sub["anomaly_flag"].sum()) if len(sub) and "anomaly_flag" in sub.columns else None,
        })
    by_type = (
        scored.groupby(["risk_tier", "type"], dropna=False)
        .size()
        .reset_index(name="count")
        if "risk_tier" in scored.columns and "type" in scored.columns
        else pd.DataFrame()
    )
    if not by_type.empty:
        by_type.columns = ["risk_tier", "transaction_type", "count"]
        return pd.DataFrame(rows), by_type
    return pd.DataFrame(rows), pd.DataFrame()


def _curated_alerts(df: pd.DataFrame) -> pd.DataFrame:
    keep = [
        "transaction_id", "risk_score", "risk_tier", "anomaly_flag",
        "type", "amount", "country", "hour",
        "nameOrig", "nameDest",
        "oldbalanceOrg", "newbalanceOrig",
        "balance_drained", "is_late_night", "is_high_risk_country", "is_large_transaction",
    ]
    if "isFraud" in df.columns:
        keep.append("isFraud")
    cols = [c for c in keep if c in df.columns]
    out = df[cols].sort_values("risk_score", ascending=False).copy()
    out.insert(0, "analyst_review_status", "Pending")
    out["analyst_notes"] = ""
    return out


def build_workpaper(output_path: Path | None = None) -> Path:
    """Build the analyst Excel workpaper from pipeline outputs."""
    output_path = output_path or WORKPAPER_PATH
    config.ensure_dirs()

    scored = _load_optional_csv(config.SCORED_CSV)
    if scored is None:
        raise FileNotFoundError(
            f"Scored transactions not found at {config.SCORED_CSV}. "
            "Run: python anomaly_detection.py  (or python run_pipeline.py)"
        )

    raw = _load_optional_csv(config.TRANSACTIONS_CSV)
    narratives = _load_optional_csv(config.SAR_NARRATIVES_CSV)
    register = _load_optional_csv(config.RISK_REGISTER_CSV)
    metrics = _metrics(scored, raw)
    date_range = _run_date_range(scored, raw)
    insights = _conclusion_insights(metrics, scored)

    paths = {
        "scored": str(config.SCORED_CSV),
        "raw": str(config.TRANSACTIONS_CSV) if raw is not None else "Not available",
        "narratives": str(config.SAR_NARRATIVES_CSV) if narratives is not None else "Not generated",
        "topics": str(config.SAR_TOPICS_CSV) if _load_optional_csv(config.SAR_TOPICS_CSV) is not None else "Not generated",
        "register": str(config.RISK_REGISTER_CSV) if register is not None else "Not generated",
    }

    wb = Workbook()
    wb.remove(wb.active)

    profit_summary = _load_optional_csv(config.PROFIT_SUMMARY_CSV)
    profit_by_tier = _load_optional_csv(config.PROFIT_BY_TIER_CSV)
    profit_by_type = _load_optional_csv(config.PROFIT_BY_TYPE_CSV)
    profit_accounts = _load_optional_csv(config.PROFIT_TOP_ACCOUNTS_CSV)

    _write_methodology(wb, metrics, paths, date_range=date_range, insights=insights)
    _write_executive_summary(wb, metrics)
    if profit_summary is not None:
        _write_dataframe_sheet(wb, "Financial Summary", profit_summary, tier_column=None)
    if profit_by_tier is not None and not profit_by_tier.empty:
        _write_dataframe_sheet(wb, "Profit by Tier", profit_by_tier, tier_column="risk_tier")
    if profit_by_type is not None and not profit_by_type.empty:
        _write_dataframe_sheet(wb, "Profit by Type", profit_by_type, tier_column=None)
    if profit_accounts is not None and not profit_accounts.empty:
        _write_dataframe_sheet(wb, "Top Account Exposure", profit_accounts, tier_column=None)
    _write_data_dictionary(wb)

    high = scored[scored["risk_tier"] == "HIGH"] if "risk_tier" in scored.columns else scored.iloc[0:0]
    medium = scored[scored["risk_tier"] == "MEDIUM"] if "risk_tier" in scored.columns else scored.iloc[0:0]

    _write_dataframe_sheet(
        wb, "HIGH Risk Alerts", _curated_alerts(high),
        extra_columns=["analyst_review_status", "analyst_notes"],
    )
    _write_dataframe_sheet(
        wb, "MEDIUM Risk Alerts", _curated_alerts(medium),
        extra_columns=["analyst_review_status", "analyst_notes"],
    )

    if narratives is not None:
        sar_cols = [
            c for c in [
                "sar_id", "transaction_id", "risk_score", "risk_tier",
                "type", "amount", "country", "hour", "sar_narrative",
            ]
            if c in narratives.columns
        ]
        sar_df = narratives[sar_cols].copy()
        sar_df["analyst_review_status"] = "Pending"
        sar_df["analyst_notes"] = ""
        _write_dataframe_sheet(wb, "SAR Narratives", sar_df)
    else:
        ws = wb.create_sheet("SAR Narratives")
        ws["A1"] = "No SAR narratives file. Run: python sar_narrative_generator.py"

    meta = _load_topic_meta()
    if meta:
        meta_df = pd.DataFrame([{
            "z_threshold": meta.get("z_threshold"),
            "n_topics": meta.get("n_topics"),
            "stop_words": ", ".join(meta.get("stop_words") or []),
        }])
        _write_dataframe_sheet(wb, "Topic Model Settings", meta_df, tier_column=None)

    topic_summary = _load_optional_csv(config.SAR_TOPICS_CSV)
    topic_assign = _load_optional_csv(config.SAR_TOPIC_ASSIGNMENTS_CSV)
    if topic_summary is not None and not topic_summary.empty:
        _write_dataframe_sheet(wb, "SAR Topic Summary", topic_summary, tier_column=None)
    else:
        ws = wb.create_sheet("SAR Topic Summary")
        ws["A1"] = "No topic summary. Save topic model from SAR narratives tab in dashboard."
    if topic_assign is not None and not topic_assign.empty:
        _write_dataframe_sheet(
            wb, "SAR Topic Assignments", topic_assign,
            tier_column="risk_tier" if "risk_tier" in topic_assign.columns else None,
        )

    word_freq = _load_optional_csv(config.SAR_WORD_FREQ_CSV)
    if word_freq is not None and not word_freq.empty:
        _write_dataframe_sheet(wb, "SAR Word Tree", word_freq, tier_column=None)
    else:
        ws = wb.create_sheet("SAR Word Tree")
        ws["A1"] = "Save topic model from dashboard SAR narratives tab to populate word frequencies."

    if register is not None:
        reg = register.copy()
        reg["analyst_review_status"] = "Pending"
        reg["analyst_notes"] = ""
        reg_cols = [
            c for c in [
                "sar_id", "transaction_id", "risk_score", "risk_tier",
                "transaction_type", "amount", "country",
                "extracted_accounts", "extracted_amounts", "extracted_countries",
                "extracted_risk_factors", "amount_reconciled", "country_reconciled",
                "extraction_method", "sar_narrative",
            ]
            if c in reg.columns
        ]
        reg = reg[reg_cols]
        if "risk_score" in reg.columns:
            reg = reg.sort_values("risk_score", ascending=False)
        _write_dataframe_sheet(
            wb, "Risk Register", reg,
            tier_column="risk_tier" if "risk_tier" in reg.columns else None,
        )
    else:
        ws = wb.create_sheet("Risk Register")
        ws["A1"] = "No risk register yet. Run: python gliner_extraction.py"

    indicators = _indicator_frame(
        pd.concat([high, medium], ignore_index=True) if len(high) + len(medium) else scored.head(0)
    )
    _write_dataframe_sheet(wb, "Risk Indicators", indicators)

    pop, pop_type = _population_summary(scored)
    _write_dataframe_sheet(wb, "Population by Tier", pop, tier_column="risk_tier")
    if not pop_type.empty:
        _write_dataframe_sheet(wb, "Population by Type", pop_type, tier_column="risk_tier")

    tmp_path = output_path.with_suffix(".tmp.xlsx")
    try:
        wb.save(tmp_path)
        if output_path.exists():
            output_path.unlink()
        shutil.move(str(tmp_path), str(output_path))
    except PermissionError as exc:
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass
        raise PermissionError(
            f"Could not write {output_path}. Close the workbook in Excel and try again."
        ) from exc
    except Exception:
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass
        raise

    print(f"Workpaper saved to {output_path}")
    return output_path


if __name__ == "__main__":
    build_workpaper()
